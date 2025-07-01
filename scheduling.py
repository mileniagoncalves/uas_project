import pandas as pd
from datetime import datetime, timedelta, time
from collections import defaultdict
import re
import openpyxl
from openpyxl.styles import PatternFill

print("\n🗕️ Running the main scheduling...")

# ================== Initial Configuration ==================
AVAILABLE_ROOMS = {
    'GD A': {
        2: [f"A2-{i}" for i in range(1, 9)],
        3: [f"A3-{i}" for i in range(1, 9)],
        4: [f"A4-{i}" for i in range(1, 9)],
        5: [f"A5-{i}" for i in range(1, 9)],
    },
    'GD B': {
        3: [f"B3-{i}" for i in range(1, 6)],
        4: [f"B4-{i}" for i in range(1, 6)],
        5: [f"B5-{i}" for i in range(1, 6)],
    }
}

ROOM_PREFERENCES = {
    'TI': {'floors': [3, 4]},
    'SI': {'floors': [3, 4]},
    'DK': {'floors': [4, 5]},
    'SD': {'floors': [2, 3]},
    'HK': {'floors': [3, 4]},
    'ME': {'floors': [4, 5]},
    'EL': {'floors': [4, 5]},
    'AKT': {'floors': [2, 3]},
    'MJN': {'floors': [2, 3]},
}

# ================== Read and Prepare Data ==================
df = pd.read_excel("teaching_data_clean.xlsx")
df["Available Day"] = df["Available Day"].fillna("ALL").str.upper()
df["Available Times"] = df["Available Times"].fillna("ALL").astype(str).str.upper()
df["Kelas"] = df["Kelas"].fillna("").astype(str).str.upper()


DAYS = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
BREAK_TIMES = [(time(12, 0), time(13, 0)), (time(18, 0), time(18, 30))]

def sks_to_duration(sks): return timedelta(minutes=50 * sks)
def is_evening_class(cls): return "M" in cls
def get_allowed_days(cls):
    if "B" in cls: return ["SATURDAY"]
    elif "C" in cls: return ["SUNDAY"]
    else: return DAYS

def get_time_window(cls):
    if is_evening_class(cls):
        return (time(17, 0), time(22, 0))
    elif "B" in cls or "C" in cls:
        return (time(8, 0), time(21, 0))
    else:
        return (time(8, 0), time(18, 0))

def generate_slots(start, end, step=10):
    slots, t = [], datetime.combine(datetime.today(), start)
    while t.time() < end:
        slots.append(t.time())
        t += timedelta(minutes=step)
    return slots

def is_in_break(start, end):
    return any(start < b_end and end > b_start for b_start, b_end in BREAK_TIMES)

def is_conflict(schedule, day, entity, start, end):
    return any(not (end <= s or start >= e) for s, e in schedule[day][entity])

def find_room(day, start, end, cls, schedule):
    match = re.match(r"([A-Z]{2})", cls)
    if not match:
        return None
    code = match.group(1)
    preferences = ROOM_PREFERENCES.get(code, {"floors": []})["floors"]
    for building, floor_dict in AVAILABLE_ROOMS.items():
        for floor in preferences:
            for room in floor_dict.get(floor, []):
                if not is_conflict(schedule, day, room, start, end):
                    return room
    return None

def find_slot(day, cls, lecturer, duration, allowed_times, schedule):
    start_win, end_win = get_time_window(cls)
    for start in generate_slots(start_win, end_win):
        end = (datetime.combine(datetime.today(), start) + duration).time()
        if end > end_win or is_in_break(start, end): continue

        if allowed_times != "ALL":
            try:
                a_start_str = allowed_times.split("-")[0].strip()
                if ":" not in a_start_str: raise ValueError("Invalid time format")
                a_start = datetime.strptime(a_start_str, "%H:%M").time()
                if start < a_start: continue
            except Exception as e:
                print(f"[⚠] Failed to parse Available Times '{allowed_times}' → {e}")
                pass

        if is_conflict(schedule, day, lecturer, start, end): continue
        if is_conflict(schedule, day, cls, start, end): continue

        return start, end
    return None, None

# ================== Scheduling ==================
final_schedule = []
failed_schedule = []
occupied_times = defaultdict(lambda: defaultdict(list))

for _, row in df.iterrows():
    lecturer, course, sks = row["DOSEN"], row["Mata Kuliah"], int(row["SKS"])
    class_list = row["Kelas"].split(",")
    available_days = row["Available Day"].split(",") if row["Available Day"] != "ALL" else DAYS
    available_times = row["Available Times"]
    duration = sks_to_duration(sks)

    for cls in class_list:
        cls = cls.strip()
        possible_days = [d for d in available_days if d in get_allowed_days(cls)]
        scheduled = False
        for day in possible_days:
            start, end = find_slot(day, cls, lecturer, duration, available_times, occupied_times)
            if start and end:
                room = find_room(day, start, end, cls, occupied_times)
                if not room:
                    continue

                today_class_schedule = [
                    j for j in final_schedule
                    if j["Kelas"] == cls and j["Hari"] == day and j["Status"] in ["SCHEDULED", "ONLINE"]
                ]

                if any(k in cls for k in ["B", "C", "M"]):
                    if len(today_class_schedule) >= 10:
                        continue
                else:
                    if len(today_class_schedule) >= 3:
                        continue

                status = "ONLINE" if end > time(21, 0) else "SCHEDULED"

                final_schedule.append({
                    "Lecturer": lecturer, "Course": course, "Class": cls,
                    "Day": day, "Time": f"{start.strftime('%H:%M')} - {end.strftime('%H:%M')}",
                    "Room": room, "Date": pd.NaT, "Status": status
                })
                occupied_times[day][lecturer].append((start, end))
                occupied_times[day][cls].append((start, end))
                occupied_times[day][room].append((start, end))
                scheduled = True
                break

        if not scheduled:
            start_win, _ = get_time_window(cls)
            start_time = start_win
            end_time = (datetime.combine(datetime.today(), start_time) + duration).time()
            final_schedule.append({
                "Lecturer": lecturer, "Course": course, "Class": cls,
                "Day": "ONLINE", "Time": f"{start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}",
                "Room": "-", "Date": pd.NaT, "Status": "ONLINE"
            })
            failed_schedule.append({
                "Lecturer": lecturer, "Course": course, "Class": cls,
                "Reason": "No available slot found for SKS and lecturer time",
                "Available Day": ", ".join(available_days),
                "Available Times": available_times,
                "SKS": sks
            })

# ================== Save Output ==================
def extract_major_batch(cls):
    match = re.match(r"([A-Z]{2})(\\d{2})", str(cls).upper())
    if match:
        return match.group(1) + "20" + match.group(2)
    return "OTHERS"

schedule_df = pd.DataFrame(final_schedule)
schedule_df["Sheet"] = schedule_df["Class"].apply(extract_major_batch)

with pd.ExcelWriter("schedule_output_with_fallback.xlsx", engine="openpyxl") as writer:
    for sheet_name, subdf in schedule_df.groupby("Sheet"):
        subdf.drop(columns=["Sheet"]).to_excel(writer, sheet_name=sheet_name, index=False)
        sheet = writer.book[sheet_name]
        for row_idx, row in subdf.iterrows():
            if row["Status"] == "ONLINE":
                for col_idx in range(1, len(subdf.columns) + 1):
                    sheet.cell(row=row_idx + 2, column=col_idx).fill = PatternFill("solid", fgColor="FFFF00")

    def clean_sheet_name(name):
        return re.sub(r"[\\/*?:\\[\\]]", "_", name)[:31]

    for lecturer_name, subdf in schedule_df.groupby("Lecturer"):
        safe_name = clean_sheet_name(lecturer_name)
        subdf.to_excel(writer, sheet_name=safe_name, index=False)

    summary_courses = schedule_df.groupby("Class")["Course"].unique().reset_index()
    summary_courses["Course"] = summary_courses["Course"].apply(
        lambda x: ", ".join(sorted({str(i) for i in x if pd.notna(i)}))
    )
    summary_courses.to_excel(writer, sheet_name="SUMMARY_COURSE_PER_CLASS", index=False)
    pd.DataFrame(failed_schedule).to_excel(writer, sheet_name="FAILED_SCHEDULE", index=False)

print("✅ Schedule successfully saved to schedule_output_with_fallback.xlsx")

